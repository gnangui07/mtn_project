"""
Vues d'export Excel pour les bons de commande
Contient toutes les fonctions d'export volumineuses (> 200 lignes)

Supporte deux modes:
- Synchrone (par défaut): Export direct, réponse bloquante
- Asynchrone (?async=1): Démarrage tâche Celery, polling via API
"""

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect, get_object_or_404
from django.contrib import messages
from django.db.models import Sum, F, Value, DecimalField, ExpressionWrapper, Case, When, Subquery, OuterRef, FloatField
import pandas as pd
from io import BytesIO
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import logging
import os
from datetime import datetime, date
from decimal import Decimal
from .views import filter_bons_by_user_service

logger = logging.getLogger(__name__)

# Import Celery tasks avec fallback si non disponible
try:
    from .tasks import export_po_progress_task, export_vendor_evaluations_task, export_bon_excel_task
    from .task_status_api import register_user_task
    CELERY_AVAILABLE = True
except ImportError:
    CELERY_AVAILABLE = False


# ============================================================================
# UTILS D'EXPORT
# ============================================================================

def normalize_header(header: str) -> str:
    """Normalise un en-tête: strip -> lower -> remplace espaces/_/- par _ -> retire accents -> compresse underscores"""
    import re
    import unicodedata
    if not header:
        return ""
    # Convertir en minuscules et retirer les espaces en début/fin
    header = str(header).strip().lower()
    # Retirer les accents
    header = ''.join(c for c in unicodedata.normalize('NFD', header)
                  if unicodedata.category(c) != 'Mn')
    # Remplacer les espaces, tirets et underscores existants par un seul underscore
    header = re.sub(r'[\s\-_]+', '_', header)
    # Supprimer les caractères non alphanumériques (sauf underscore)
    header = re.sub(r'[^\w_]', '', header)
    # Supprimer les underscores au début ou à la fin
    header = header.strip('_')
    return header

def get_value_tolerant(contenu: dict, exact_candidates=None, tokens=None):
    """Retourne la valeur pour une clé en acceptant des variantes d'en-têtes.
    - exact_candidates: liste de libellés candidats (str) comparés après normalisation
    - tokens: liste de mots qui doivent tous être présents dans l'en-tête normalisé
    Gestionne notamment: espaces de fin/début, doubles espaces, underscores, casse.
    """
    if not contenu:
        return None
    
    # Construire un mapping normalisé -> (clé originale, valeur)
    normalized = {normalize_header(k): (k, v) for k, v in contenu.items() if k}

    # 1) Essais exacts (après normalisation)
    if exact_candidates:
        for cand in exact_candidates:
            nk = normalize_header(cand)
            if nk in normalized:
                val = normalized[nk][1]
                return val if val not in (None, '') else None

    # 2) Recherche par tokens (tous présents dans la clé normalisée)
    if tokens:
        needed = [normalize_header(t) for t in tokens]
        for nk, (_ok, val) in normalized.items():
            if all(t in nk for t in needed):
                return val if val not in (None, '') else None
    return None

def find_order_key(contenu: dict):
    # 1) Cas exact 'Order'
    if 'Order' in contenu:
        return 'Order'
    # 2) Recherche tolérante (espaces/underscores/accents)
    for k in contenu.keys():
        if not k:
            continue
        norm = normalize_header(k)
        if 'order' in norm or 'commande' in norm or 'bon' in norm or norm == 'bc':
            return k
    return None

def _make_export_payload(bon_id):
    """
    Fonction utilitaire pour construire le payload d'export à partir d'un bon de commande.
    Récupère les lignes de fichier associées et les informations de réception.
    """
    from .models import NumeroBonCommande, LigneFichier, Reception
    import logging
    logger = logging.getLogger(__name__)
    
    try:
        bon = NumeroBonCommande.objects.get(id=bon_id)
        fichier = bon.fichiers.first() # Prend le premier fichier associé
        if not fichier:
            return []

        lignes = LigneFichier.objects.filter(fichier=fichier).order_by('numero_ligne')
        receptions = Reception.objects.filter(bon_commande=bon).order_by('business_id')

        # Créer des dictionnaires pour un accès rapide
        receptions_dict = {r.business_id: r for r in receptions}
        
        payload = []
        for ligne in lignes:
            contenu = ligne.contenu or {}
            business_id = ligne.business_id
            
            # Récupérer les infos de réception si disponibles
            reception_info = receptions_dict.get(business_id)
            
            def format_decimal(val):
                if val is None: return None
                try:
                    return f"{Decimal(str(val)):.2f}"
                except:
                    return str(val)

            item_data = {
                'business_id': business_id,
                'order': contenu.get('Order'),
                'line': contenu.get('Line'),
                'item': contenu.get('Item'),
                'schedule': contenu.get('Schedule'),
                'ordered_quantity': format_decimal(reception_info.ordered_quantity) if reception_info and reception_info.ordered_quantity is not None else contenu.get('Ordered Quantity'),
                'quantity_delivered': format_decimal(reception_info.quantity_delivered) if reception_info and reception_info.quantity_delivered is not None else contenu.get('Quantity Delivered'),
                'received_quantity': format_decimal(reception_info.received_quantity) if reception_info and reception_info.received_quantity is not None else contenu.get('Received Quantity'),
                'quantity_not_delivered': format_decimal(reception_info.quantity_not_delivered) if reception_info and reception_info.quantity_not_delivered is not None else contenu.get('Quantity Not Delivered'),
                'unit_price': format_decimal(reception_info.unit_price) if reception_info and reception_info.unit_price is not None else contenu.get('Price'),
                'amount_delivered': format_decimal(reception_info.amount_delivered) if reception_info and reception_info.amount_delivered is not None else contenu.get('Amount Delivered'),
                'currency': contenu.get('Currency'),
                'price': contenu.get('Price'), # Garder le prix original du fichier
                'description': contenu.get('Description'),
                'project_number': contenu.get('Project Number'),
                'cpu': contenu.get('CPU'),
                'sponsor': contenu.get('Sponsor'),
                'project_manager': contenu.get('Project Manager'),
                'order_description': contenu.get('Order Description'),
                'asset_type': contenu.get('ASSET TYPE'),
                'pip_end_date': contenu.get('PIP END DATE'),
                'revised_end_date': contenu.get('REVISED END DATE'),
                'actual_end_date': contenu.get('ACTUAL END DATE'),
                'line_type': contenu.get('Line Type'),
                'code_ifs': contenu.get('Code - IFS'),
                # Retrait des champs inexistants sur Reception
                'irb_id': str(reception_info.id) if reception_info else None,
            }
            payload.append(item_data)
            
        return payload

    except NumeroBonCommande.DoesNotExist:
        return []
    except Exception as e:
        # Log the error for debugging
        logger.error(f"Error in _make_export_payload for bon_id {bon_id}: {e}")
        return []

# ============================================================================
# EXPORTS EXCEL
# ============================================================================

@login_required
def export_po_progress_monitoring(request):
    """
    But:
    - Exporter en Excel un tableau récapitulatif des bons (1 ligne par PO) pour le suivi.

    Mode async:
    - Si ?async=1 : Lance la tâche Celery et retourne un JSON avec task_id pour polling
    - Sinon : Export synchrone classique

    Entrées:
    - request (HttpRequest): utilisateur connecté requis (login_required).
    - GET param 'async' (optionnel): si '1', mode asynchrone

    Sorties:
    - Si async=1: JsonResponse avec task_id
    - Sinon: HttpResponse fichier .xlsx en téléchargement
    """
    # ===== MODE ASYNC =====
    if request.GET.get('async') == '1' and CELERY_AVAILABLE:
        try:
            # Récupérer le service de l'utilisateur pour filtrer
            user_service = None
            if hasattr(request.user, 'service') and request.user.service:
                user_service = request.user.service
            
            # Lancer la tâche Celery
            task = export_po_progress_task.delay(
                user_id=request.user.id,
                user_service=user_service
            )
            
            # Enregistrer la tâche pour l'utilisateur
            try:
                register_user_task(request.user.id, task.id, 'export_po_progress')
            except Exception:
                pass
            
            return JsonResponse({
                'success': True,
                'async': True,
                'task_id': task.id,
                'message': 'Export démarré en arrière-plan',
                'poll_url': f'/orders/api/task-status/{task.id}/'
            })
        except Exception as e:
            logger.error(f"Erreur lors du démarrage de l'export async: {e}")
            # Fallback vers le mode sync si erreur
            pass
    
    # ===== MODE SYNC (par défaut) =====
    try:
        from .models import NumeroBonCommande, InitialReceptionBusiness, LigneFichier

        # Récupération des données des fichiers liés aux bons de commande
        bons_commande_qs = NumeroBonCommande.objects.prefetch_related(
            'fichiers__lignes',
            'receptions'
        ).all()

        # Filtrer par service/CPU (Option A) pour les non-superusers
        if not request.user.is_superuser:
            from .views import filter_bons_by_user_service
            bons_commande_qs = filter_bons_by_user_service(bons_commande_qs, request.user)

        bons_commande = bons_commande_qs
        
        # Debug: Afficher les clés disponibles dans la première occurrence
        debug_first_occurrence = None
        
        # Étape 1: Créer un cache pour les premières occurrences de chaque bon de commande en une seule passe globale
        orders_needed = {bon.numero for bon in bons_commande}

        # Lister tous les fichiers associés, triés par date d'importation
        all_files = []
        for bon in bons_commande:
            for fichier in bon.fichiers.all():
                all_files.append(fichier)
        all_files.sort(key=lambda f: f.date_importation)

        first_occurrence_cache = {}
        for fichier in all_files:
            lignes = list(fichier.lignes.all())
            lignes.sort(key=lambda l: l.numero_ligne)
            for ligne in lignes:
                contenu = ligne.contenu or {}
                order_key = find_order_key(contenu)
                if not order_key:
                    continue
                order_val = str(contenu.get(order_key, '')).strip()
                if not order_val:
                    continue
                if order_val in orders_needed and order_val not in first_occurrence_cache:
                    first_occurrence_cache[order_val] = contenu
                    if len(first_occurrence_cache) == len(orders_needed):
                        break
            if len(first_occurrence_cache) == len(orders_needed):
                break

        # Étape 2: Préparation des données pour le DataFrame
        data = []
        for bon in bons_commande:
            # Récupérer la première occurrence de ce bon de commande
            premiere_occurrence = first_occurrence_cache.get(bon.numero)
            
            # Récupérer l'évaluation fournisseur depuis la base de données
            from .models import VendorEvaluation, TimelineDelay
            vendor_evaluation = VendorEvaluation.objects.filter(bon_commande=bon).first()
            
            # Récupérer les commentaires Timeline Delays
            timeline_delay = TimelineDelay.objects.filter(bon_commande=bon).first()
            
            # Debug: Stocker la première occurrence pour inspection
            if debug_first_occurrence is None and premiere_occurrence:
                debug_first_occurrence = premiere_occurrence
                print("Debug - Clés disponibles dans la première occurrence:", sorted(premiere_occurrence.keys()))

            # Si aucune occurrence n'a été trouvée, passer à la suite
            if premiere_occurrence is None:
                print(f"Avertissement: Aucune occurrence trouvée pour le bon de commande {bon.numero}")
                continue

            # Récupération des montants avec conversion en float pour éviter les problèmes de type
            from decimal import Decimal
            montants = InitialReceptionBusiness.objects.filter(
                bon_commande=bon
            ).aggregate(
                total_recu=Sum('montant_recu_initial'),
                total_initial=Sum('montant_total_initial')
            )

            # Conversion des valeurs décimales en float
            montants = {k: float(v if v is not None else 0) for k, v in montants.items()}

            # Calcul du pourcentage financier
            total_initial = montants['total_initial'] or 0
            total_recu = montants['total_recu'] or 0
            financial_percent = ((total_recu / total_initial) * 100) if total_initial > 0 else 0
            currency = premiere_occurrence.get('Currency')
            if currency == 'XOF':
                exchange_rate = 1.0
            else:
                # Récupérer le taux de conversion de la première occurrence
                exchange_rate_str = premiere_occurrence.get('Conversion Rate')
                try:
                    exchange_rate = float(exchange_rate_str) if exchange_rate_str else 1.0
                except (TypeError, ValueError):
                    exchange_rate = 1.0

            # Calcul du PO en XOF
            po_amount = float(bon.montant_total()) if callable(getattr(bon, 'montant_total', None)) else 0.0
            po_amount_xof = po_amount * exchange_rate


            # Recalculer le retard total à partir des dates (supporte plusieurs formats)
            from datetime import datetime
            pip_end_str = premiere_occurrence.get('PIP END DATE')
            actual_end_str = premiere_occurrence.get('ACTUAL END DATE')
            total_days_late = 0
            if pip_end_str and actual_end_str:
                try:
                    # Convertir en string et nettoyer
                    pip_end_str = str(pip_end_str).strip()
                    actual_end_str = str(actual_end_str).strip()
                    
                    # Liste des formats supportés
                    date_formats = [
                        '%Y-%m-%d %H:%M:%S',  # 2025-07-30 00:00:00
                        '%Y-%m-%d',           # 2025-07-30
                        '%d/%m/%Y',           # 30/07/2025
                        '%d/%m/%Y %H:%M:%S'   # 30/07/2025 00:00:00
                    ]
                    
                    pip_end = None
                    actual_end = None
                    
                    # Essayer chaque format pour PIP END DATE
                    for fmt in date_formats:
                        try:
                            pip_end = datetime.strptime(pip_end_str, fmt)
                            break
                        except ValueError:
                            continue
                    
                    # Essayer chaque format pour ACTUAL END DATE
                    for fmt in date_formats:
                        try:
                            actual_end = datetime.strptime(actual_end_str, fmt)
                            break
                        except ValueError:
                            continue
                    
                    if pip_end and actual_end:
                        total_days_late = (actual_end - pip_end).days
                    else:
                        total_days_late = 0
                except Exception:
                    total_days_late = 0

            # Calcul du retard imputable au vendeur à partir du total et de la force majeure
            force_majeure_value = premiere_occurrence.get('Day Late Due to Force Majeure')
            if force_majeure_value is None:
                force_majeure_value = 0
            try:
                day_late_force_majeure = float(force_majeure_value)
            except (TypeError, ValueError):
                day_late_force_majeure = 0

            day_late_due_to_vendor = max(0, total_days_late - day_late_force_majeure)

            # Calcul de Accruals (Cur)
            montant_recu = float(bon.montant_recu()) if callable(getattr(bon, 'montant_recu', None)) else 0.0
            accruals_cur = montant_recu - float(total_recu)

            # Calcul de Receipt Not Invoiced (CUR)
            invoiced_amount_str = premiere_occurrence.get("Invoiced Amount")
            try:
                invoiced_amount = float(invoiced_amount_str) if invoiced_amount_str else 0.0
            except (TypeError, ValueError):
                invoiced_amount = 0.0
            receipt_not_invoiced = float(total_recu) - invoiced_amount

            # Calcul des nouvelles colonnes
            po_amount = float(bon.montant_total()) if callable(getattr(bon, 'montant_total', None)) else 0.0
            
            # 1. Calcul initial de Penalties (0.3% * Day Late due to Vendor * PO Amount)
            penalties = 0.003 * day_late_due_to_vendor * po_amount
            
            # 2. Calcul de Rate PO Amount (PO Amount * 10%)
            rate_po_amount = po_amount * 0.10
            
            # 3. Application de la condition : 
            # Si penalties > rate_po_amount, on prend rate_po_amount, sinon on garde penalties
            if penalties > rate_po_amount:
                penalties = rate_po_amount

            # Pré-calculs pour limiter les appels et accélérer
            taux_avancement_percent = float(bon.taux_avancement()) if hasattr(bon, 'taux_avancement') and callable(bon.taux_avancement) else 0.0
            current_onground = taux_avancement_percent  # Garder en pourcentage pour l'affichage
            current_onground_fraction = taux_avancement_percent / 100  # Fraction pour les calculs
            delivery_amount_cur = current_onground_fraction * po_amount
            delivery_amount_xof = current_onground_fraction * po_amount_xof
            financial_percent_float = float(financial_percent)
            receipt_amount = float(total_recu)
            invoiced_amount_value = int(float(premiere_occurrence.get("Invoiced Amount", 0))) if premiere_occurrence.get("Invoiced Amount") else 0
            retention_rate_fraction = float(bon.retention_rate) if bon.retention_rate is not None else 0.0
            pip_retention_percent_val = ((penalties / po_amount) * 100) if po_amount > 0 else 0.0
            total_retention_val = pip_retention_percent_val + retention_rate_fraction

            # Préparation de la ligne de données
            data.append({
                'numero': bon.numero,
                'cpu': premiere_occurrence.get('CPU'),
                'sponsor': premiere_occurrence.get('Sponsor'),
                'project_number': premiere_occurrence.get('Project Number'),
                'project_manager': premiere_occurrence.get(' Project Manager') or premiere_occurrence.get('Project Manager') or premiere_occurrence.get('Project Manager Name') or premiere_occurrence.get('Project_Manager') or premiere_occurrence.get('ProjectManager') or premiere_occurrence.get('PM') or '',
                'project_coordinator': premiere_occurrence.get('Project Coordinator'),
                'senior_technical_lead': premiere_occurrence.get('Senior Technical Lead'),
                'supplier': premiere_occurrence.get('Supplier'),
                'order_description': premiere_occurrence.get('Order Description'),
                'currency': premiere_occurrence.get('Currency'),
                'annee': premiere_occurrence.get('Année'),
                'project_name': premiere_occurrence.get('Project Name'),
                'po_type': premiere_occurrence.get('Po type'),
                'replaced_order': premiere_occurrence.get('Replaced Order'),
                'asset_type': premiere_occurrence.get('ASSET TYPE'),
                'pip_end_date': premiere_occurrence.get('PIP END DATE'),
                'revised_end_date': premiere_occurrence.get('REVISED END DATE'),
                'actual_end_date': premiere_occurrence.get('ACTUAL END DATE'),
                'line_type': premiere_occurrence.get('Line Type'),
                'code_ifs': premiere_occurrence.get('Code - IFS'),
                'po_amount': po_amount,
                'exchange_rate': exchange_rate,
                'po_amount_xof': po_amount_xof,
                'current_onground_percent': current_onground,
                'financial_percent': financial_percent_float,  # Convertir en décimal pour le format pourcentage
                'delivery_amount_cur': delivery_amount_cur,
                'delivery_amount_xof': delivery_amount_xof,
                'receipt_amount': receipt_amount,
                'invoiced_amount': invoiced_amount_value,
                'receipt_not_invoiced': receipt_not_invoiced,
                'accruals_cur': accruals_cur,
                'day_late_due_to_mtn': premiere_occurrence.get('Day Late Due to MTN'),
                'day_late_due_to_force_majeure': premiere_occurrence.get('Day Late Due to Force Majeure'),
                'total_days_late': total_days_late,
                'day_late_due_to_vendor': day_late_due_to_vendor,
                'retention_cause': bon.retention_cause or '',
                'penalties': penalties,  # Colonne Penalties (déjà mise à jour avec la condition)
                'pip_retention_percent': pip_retention_percent_val,  # % PIP retention
                'other_retentions': retention_rate_fraction,
                'total_retention': total_retention_val,
                # Colonnes d'évaluation fournisseur: Priorité BD, fallback fichier Excel
                'Delivery_Compliance_To_Order': (
                    vendor_evaluation.delivery_compliance if vendor_evaluation else
                    get_value_tolerant(premiere_occurrence, exact_candidates=['Delivery Compliance to Order (Quantity & Quality)', 'Delivery Compliance to Order', 'Delivery Compliance To Order'], tokens=['delivery', 'compliance', 'order']) or ''
                ),
                'Delivery_Execution_Timeline': (
                    vendor_evaluation.delivery_timeline if vendor_evaluation else
                    get_value_tolerant(premiere_occurrence, exact_candidates=['Delivery Execution  Timeline', 'Delivery Execution Timeline'], tokens=['delivery', 'execution', 'timeline']) or ''
                ),
                'Vendor_Advising_Capability': (
                    vendor_evaluation.advising_capability if vendor_evaluation else
                    get_value_tolerant(premiere_occurrence, exact_candidates=['Vendor Advising Capability'], tokens=['vendor', 'advising', 'capability']) or ''
                ),
                'After_Sales_Services_QOS': (
                    vendor_evaluation.after_sales_qos if vendor_evaluation else
                    get_value_tolerant(premiere_occurrence, exact_candidates=['After Sales Services QOS', 'After Sales Services QoS', 'After Sales Service QOS'], tokens=['after', 'sales', 'services', 'qos']) or ''
                ),
                'Vendor_Relationship': (
                    vendor_evaluation.vendor_relationship if vendor_evaluation else
                    get_value_tolerant(premiere_occurrence, exact_candidates=['Vendor Relationship'], tokens=['vendor', 'relationship']) or ''
                ),
                'Vendor_Final_Rating': (
                    float(vendor_evaluation.vendor_final_rating) if vendor_evaluation else
                    get_value_tolerant(premiere_occurrence, exact_candidates=['Vendor Final Rating'], tokens=['vendor', 'final', 'rating']) or ''
                ),
                # Commentaires Timeline Delays
                'comment_mtn': timeline_delay.comment_mtn if timeline_delay else '',
                'comment_force_majeure': timeline_delay.comment_force_majeure if timeline_delay else '',
                'comment_vendor': timeline_delay.comment_vendor if timeline_delay else '',
            })

        # Création du DataFrame à partir des données préparées
        df = pd.DataFrame(data)

        # Renommage des colonnes pour l'export
        column_mapping = {
            'numero': 'Order Number',
            'cpu': 'CPU',
            'sponsor': 'Sponsor',
            'project_number': 'Project Number',
            'project_manager': 'Project Manager',
            'project_coordinator': 'Project Coordinator',
            'senior_technical_lead': 'Senior Technical Lead',
            'supplier': 'Supplier Name',
            'order_description': 'Order Description',
            'currency': 'Currency',
            'annee': 'Année',
            'project_name': 'Project Name',
            'po_type': 'PO Type',
            'replaced_order': 'Replaced Order',
            'asset_type': 'Asset Type',
            'pip_end_date': 'PIP End Date',
            'revised_end_date': 'Revised End Date',
            'actual_end_date': 'Actual End Date',
            'line_type': 'Line Type',
            'code_ifs': 'Code-IFS',
            'po_amount': 'PO Amount',
            'exchange_rate': 'Exchange Rate',
            'po_amount_xof': 'PO XOF',
            'current_onground_percent': 'Current Onground (%)',
            'financial_percent': '%Financial',
            'delivery_amount_cur': 'Delivery Amount (CUR)',
            'delivery_amount_xof': 'Delivery Amount (XOF)',
            'receipt_amount': 'Receipt Amount',
            'invoiced_amount': 'Invoiced Amount',
            'receipt_not_invoiced': 'Receipt Not Invoiced (CUR)',
            'accruals_cur': 'Accruals (Cur)',
            'day_late_due_to_mtn': 'Day Late Due to MTN',
            'day_late_due_to_force_majeure': 'Day Late Due to Force Majeure',
            'total_days_late': '# Total Days Late',
            'day_late_due_to_vendor': 'Day Late Due to Vendor',
            'retention_cause': 'Cause rétention',
            'penalties': 'Penalties',
            'pip_retention_percent': '% PIP Retention',
            'other_retentions': 'Other Retentions (%)',
            'total_retention': 'Total Retention (%)',
            'Delivery_Compliance_To_Order': 'Delivery Compliance To Order',
            'Delivery_Execution_Timeline': 'Delivery Execution Timeline',
            'Vendor_Advising_Capability': 'Vendor Advising Capability',
            'After_Sales_Services_QOS': 'After Sales Services QOS',
            'Vendor_Relationship': 'Vendor Relationship',
            'Vendor_Final_Rating': 'Vendor Final Rating',
            'comment_mtn': 'Commentaire Part MTN',
            'comment_force_majeure': 'Commentaire Part Force Majeure',
            'comment_vendor': 'Commentaire Part Fournisseur',
        }
        df.rename(columns=column_mapping, inplace=True)

        # Normaliser les colonnes de dates en chaînes 'YYYY-MM-DD' (format base de données)
        from datetime import date, datetime as dt
        for _col in ['PIP End Date', 'Revised End Date', 'Actual End Date']:
            if _col in df.columns:
                def _to_iso(v):
                    if v is None:
                        return ''
                    if isinstance(v, (dt, date)):
                        return v.strftime('%Y-%m-%d')
                    if isinstance(v, str):
                        s = v.strip()
                        # Déjà au format ISO
                        if len(s) >= 10 and s[4] == '-' and s[7] == '-':
                            return s[:10]
                        # Essayer dd/mm/YYYY
                        try:
                            return dt.strptime(s, '%d/%m/%Y').strftime('%Y-%m-%d')
                        except Exception:
                            pass
                        # Essayer dd-mm-YYYY
                        try:
                            return dt.strptime(s, '%d-%m-%Y').strftime('%Y-%m-%d')
                        except Exception:
                            pass
                        # Par défaut, renvoyer tel quel
                        return s
                    return str(v)
                df[_col] = df[_col].map(_to_iso)

        # Forcer l'affichage des montants sans décimales pour Invoiced Amount
        if 'Invoiced Amount' in df.columns:
            df['Invoiced Amount'] = df['Invoiced Amount'].astype(int)

        # Création du fichier Excel avec mise en forme
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='PO Progress Monitoring')
            workbook = writer.book
            worksheet = writer.sheets['PO Progress Monitoring']

            # Styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Format header
            for col_num, value in enumerate(df.columns.values, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment
                cell.border = thin_border
                worksheet.column_dimensions[get_column_letter(col_num)].width = max(len(value) + 2, 15)

            # Apply borders to data cells (no per-cell numeric formatting to improve speed)
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row,
                                           min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.border = thin_border

            # Format only specific numeric columns to reduce processing time
            numeric_cols = [
                'PO Amount', 'PO XOF', 'Receipt Amount', 'Accruals (Cur)',
                'Receipt Not Invoiced (CUR)', 'Delivery Amount (CUR)', 'Delivery Amount (XOF)', 'Penalties'
            ]
            for col_name in numeric_cols:
                if col_name in df.columns:
                    col_idx = df.columns.get_loc(col_name) + 1
                    for row in range(2, worksheet.max_row + 1):
                        cell = worksheet.cell(row=row, column=col_idx)
                        cell.number_format = '0.00'

            # Format percentages
            # Note: Les valeurs sont déjà en pourcentage (10 pour 10%), format français avec virgule
            percent_columns = ['Current Onground (%)', 'Retention Rate (%)', '%Financial', '% PIP Retention', 'Other Retentions (%)', 'Total Retention (%)']
            for col_name in percent_columns:
                if col_name in df.columns:
                    col_idx = df.columns.get_loc(col_name) + 1
                    for row in range(2, worksheet.max_row + 1):
                        cell = worksheet.cell(row=row, column=col_idx)
                        if cell.value is not None:
                            # Formater avec virgule et supprimer les zéros inutiles
                            formatted = f"{float(cell.value):.2f}".replace('.', ',').rstrip('0').rstrip(',')
                            cell.value = f"{formatted}%"
                        else:
                            cell.value = ""
                        cell.number_format = '@'  # Format texte pour garder le %

            # Force text format for date-like columns to preserve original values exactly as in source file
            text_cols = ['PIP End Date', 'Revised End Date', 'Actual End Date']
            for col_name in text_cols:
                if col_name in df.columns:
                    col_idx = df.columns.get_loc(col_name) + 1
                    for row in range(2, worksheet.max_row + 1):
                        cell = worksheet.cell(row=row, column=col_idx)
                        cell.number_format = '@'  # Treat as text in Excel

            # Ensure Invoiced Amount shows without decimals
            if 'Invoiced Amount' in df.columns:
                inv_idx = df.columns.get_loc('Invoiced Amount') + 1
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row, column=inv_idx)
                    cell.number_format = '0'
            
            # Format des colonnes de commentaires (largeur plus grande et wrap text)
            comment_cols = ['Commentaire Part MTN', 'Commentaire Part Force Majeure', 'Commentaire Part Fournisseur']
            for col_name in comment_cols:
                if col_name in df.columns:
                    col_idx = df.columns.get_loc(col_name) + 1
                    col_letter = get_column_letter(col_idx)
                    worksheet.column_dimensions[col_letter].width = 50  # Largeur plus grande pour les commentaires
                    for row in range(2, worksheet.max_row + 1):
                        cell = worksheet.cell(row=row, column=col_idx)
                        cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')

        # Préparer la réponse HTTP
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="po_progress_monitoring.xlsx"'

        return response

    except Exception as e:
        logger.error(f"Erreur lors de l'export PO Progress Monitoring: {str(e)}")
        return HttpResponse(
            f"Une erreur s'est produite lors de l'export: {str(e)}",
            status=500
        )

@login_required
def export_msrn_po_lines(request, msrn_id):
    """
    But:
    - Exporter en Excel les lignes d’un rapport MSRN (PO Lines) avec les montants utiles.

    Étapes:
    1) Retrouver le MSRN, son PO, et charger les réceptions.
    2) Préparer un DataFrame pour les lignes (quantités, montants, payables…).
    3) Préparer un onglet d’information MSRN (en‑tête simple).
    4) Écrire les 2 onglets dans un .xlsx et le renvoyer.

    Entrées:
    - request (HttpRequest)
    - msrn_id (int): identifiant du rapport MSRN.

    Sorties:
    - HttpResponse: fichier .xlsx contenant 2 onglets (MSRN Information, Purchase Order Lines).
    """
    try:
        from .models import MSRNReport, Reception, LigneFichier
        from decimal import Decimal
        import pandas as pd
        from io import BytesIO
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
        
        # Récupérer le rapport MSRN (retourner 404 JSON si inexistant)
        try:
            msrn_report = MSRNReport.objects.get(id=msrn_id)
        except MSRNReport.DoesNotExist:
            return HttpResponse("MSRN Report not found", status=404)
        bon_commande = msrn_report.bon_commande
        retention_rate = msrn_report.retention_rate or Decimal('0')
        
        # Récupérer les réceptions depuis la base de données
        receptions = Reception.objects.filter(bon_commande=bon_commande).order_by('business_id')
        
        # Extraire les lignes de commande
        po_lines = list(receptions)
        
        if not po_lines:
            return HttpResponse("No Purchase Order Lines found for this MSRN", status=404)
        
        # Préparer les données pour le DataFrame
        data = []
        for reception in po_lines:
            # Récupérer les informations de ligne depuis le fichier
            line_description = "N/A"
            line = "N/A"
            schedule = "N/A"
            
            try:
                lf = LigneFichier.objects.filter(business_id=reception.business_id).order_by('-id').first()
                if lf and lf.contenu:
                    line_description = lf.contenu.get('Line Description', 'N/A')
                    line = lf.contenu.get('Line', 'N/A')
                    schedule = lf.contenu.get('Schedule', 'N/A')
            except Exception:
                pass
            
            # Calculer Net Qty to Receipt in Boost
            quantity_payable = reception.quantity_payable or 0
            received_quantity = reception.received_quantity or 0
            net_qty_to_receipt = Decimal(str(quantity_payable)) - Decimal(str(received_quantity))
            
            # Calculer Amount Payable avec retention
            amount_delivered = reception.amount_delivered or 0
            retention_amount = amount_delivered * (retention_rate / Decimal('100'))
            amount_payable = amount_delivered - retention_amount
            
            row = {
                'Purchase Order': bon_commande.numero if bon_commande else 'N/A',
                'Line': line,
                'Schedule': schedule,
                'Line Description': line_description,
                'Ordered Quantity': float(reception.ordered_quantity or 0),
                'Received Quantity': float(received_quantity),
                'Quantity Delivered': float(reception.quantity_delivered or 0),
                'Unit Price': float(reception.unit_price or 0),
                'Amount Delivered': float(amount_delivered),
                'Net Qty to Receipt in Boost': float(net_qty_to_receipt),
                'Quantity Payable': float(quantity_payable),
                'Amount Payable': float(amount_payable),
            }
            data.append(row)
        
        # Créer le DataFrame pour les lignes PO
        df_lines = pd.DataFrame(data)
        
        # Préparer les données du MSRN (en-tête)
        msrn_info = {
            'Field': [
                'MSRN Number',
                'Purchase Order',
                'Supplier',
                'Currency',
                'PO Amount',
                'Amount Delivered',
                'Delivery Rate (%)',
                'Payment Retention Rate (%)',
                'Retention Cause',
                'Created Date',
            ],
            'Value': [
                msrn_report.report_number,
                bon_commande.numero if bon_commande else 'N/A',
                bon_commande.get_supplier() if bon_commande and callable(getattr(bon_commande, 'get_supplier', None)) else 'N/A',
                bon_commande.get_currency() if bon_commande and callable(getattr(bon_commande, 'get_currency', None)) else 'N/A',
                float(bon_commande.montant_total()) if bon_commande and callable(getattr(bon_commande, 'montant_total', None)) else 0,
                float(bon_commande.montant_recu()) if bon_commande and callable(getattr(bon_commande, 'montant_recu', None)) else 0,
                float(msrn_report.progress_rate_snapshot or 0),
                float(retention_rate),
                msrn_report.retention_cause or 'N/A',
                msrn_report.created_at.strftime('%Y-%m-%d %H:%M:%S') if msrn_report.created_at else 'N/A',
            ]
        }
        df_msrn = pd.DataFrame(msrn_info)
        
        # Créer le fichier Excel avec mise en forme
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Onglet 1: Informations MSRN
            df_msrn.to_excel(writer, index=False, sheet_name='MSRN Information')
            
            # Onglet 2: Purchase Order Lines
            df_lines.to_excel(writer, index=False, sheet_name='Purchase Order Lines')
            
            workbook = writer.book
            
            # ===== FORMATAGE ONGLET MSRN INFORMATION =====
            ws_msrn = writer.sheets['MSRN Information']
            
            # Styles pour MSRN Info
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            field_font = Font(bold=True)
            field_fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
            alignment_center = Alignment(horizontal="center", vertical="center")
            alignment_left = Alignment(horizontal="left", vertical="center")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Format header MSRN
            for col_num in range(1, 3):
                cell = ws_msrn.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment_center
                cell.border = thin_border
            
            ws_msrn.column_dimensions['A'].width = 30
            ws_msrn.column_dimensions['B'].width = 50
            
            # Format data cells MSRN
            for row in range(2, ws_msrn.max_row + 1):
                # Colonne Field (A)
                cell_field = ws_msrn.cell(row=row, column=1)
                cell_field.font = field_font
                cell_field.fill = field_fill
                cell_field.alignment = alignment_left
                cell_field.border = thin_border
                
                # Colonne Value (B)
                cell_value = ws_msrn.cell(row=row, column=2)
                cell_value.alignment = alignment_left
                cell_value.border = thin_border
                
                # Format numérique si applicable
                if isinstance(cell_value.value, (int, float)):
                    cell_value.number_format = '#,##0.00'
            
            # ===== FORMATAGE ONGLET PURCHASE ORDER LINES =====
            worksheet = writer.sheets['Purchase Order Lines']
            
            # Styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            alignment_center = Alignment(horizontal="center", vertical="center")
            alignment_right = Alignment(horizontal="right", vertical="center")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Format header
            for col_num, value in enumerate(df_lines.columns.values, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment_center
                cell.border = thin_border
                worksheet.column_dimensions[get_column_letter(col_num)].width = max(len(str(value)) + 2, 15)
            
            # Format data cells
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row,
                                           min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.border = thin_border
                    # Aligner les nombres à droite
                    if isinstance(cell.value, (int, float)):
                        cell.alignment = alignment_right
                        cell.number_format = '#,##0.00'
            
            # Ajouter une ligne de total si nécessaire
            total_row = worksheet.max_row + 2
            worksheet.cell(row=total_row, column=1, value="TOTAL")
            worksheet.cell(row=total_row, column=1).font = Font(bold=True)
            
        # Préparer la réponse HTTP
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"MSRN_{msrn_report.report_number}_PO_Lines.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except Exception as e:
        logger.error(f"Erreur lors de l'export des PO Lines: {str(e)}")
        return HttpResponse(
            f"Une erreur s'est produite lors de l'export: {str(e)}",
            status=500
        )

@login_required
def export_vendor_evaluations(request):
    """
    Exporte toutes les évaluations des fournisseurs en Excel avec logique robuste
    """
    from .models import VendorEvaluation, NumeroBonCommande
    import pandas as pd
    from io import BytesIO
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime
    from decimal import Decimal
    
    # Gestion du mode asynchrone
    if request.GET.get('async') == '1' and CELERY_AVAILABLE:
        try:
            task = export_vendor_evaluations_task.delay(
                user_id=request.user.id,
                filters=request.GET.dict()
            )
            # Enregistrer la tâche pour l'utilisateur (optionnel, selon votre système de polling)
            if 'register_user_task' in globals() or 'register_user_task' in locals() or globals().get('register_user_task'):
                register_user_task(request.user, task.id, "Export Évaluations Fournisseurs")
            
            return JsonResponse({
                'success': True,
                'async': True,
                'task_id': task.id
            })
        except Exception as e:
            logger.error(f"Erreur lors du lancement de l'export asynchrone: {e}")
            # Fallback sur le mode synchrone si Celery échoue
    
    # Helper: fonction pour trouver la clé 'Order' de manière tolérante
    def find_order_key(contenu: dict):
        if 'Order' in contenu:
            return 'Order'
        for k in contenu.keys():
            if not k:
                continue
            norm = ' '.join(str(k).strip().lower().replace('_', ' ').split())
            if 'order' in norm or 'commande' in norm or 'bon' in norm or norm == 'bc':
                return k
        return None
    
    # Helper: récupérer une valeur par normalisation tolérante des en-têtes
    def get_value_tolerant(contenu: dict, exact_candidates=None, tokens=None):
        if not contenu:
            return None
        def norm(s: str):
            return ' '.join(str(s).strip().lower().replace('_', ' ').replace('-', ' ').split())
        normalized = {norm(k): (k, v) for k, v in contenu.items() if k}
        
        # Essais exacts (après normalisation)
        if exact_candidates:
            for cand in exact_candidates:
                nk = norm(cand)
                if nk in normalized:
                    return normalized[nk][1]
        
        # Recherche par tokens
        if tokens:
            needed = [norm(t) for t in tokens]
            for nk, (_ok, v) in normalized.items():
                if all(t in nk for t in needed):
                    return v
        return None
    
    # Récupérer les évaluations avec filtres (mêmes que la liste)
    evaluations = VendorEvaluation.objects.select_related(
        'bon_commande', 'evaluator'
    ).prefetch_related('bon_commande__fichiers__lignes').order_by('supplier', '-date_evaluation')

    # Filtrer par service/CPU (Option A) pour les non-superusers
    if not request.user.is_superuser:
        from .views import filter_bons_by_user_service
        allowed_bons = filter_bons_by_user_service(NumeroBonCommande.objects.all(), request.user)
        evaluations = evaluations.filter(bon_commande__in=allowed_bons)
    
    # Appliquer les filtres de la requête
    supplier_filter = request.GET.get('supplier', '').strip()
    if supplier_filter:
        evaluations = evaluations.filter(supplier__icontains=supplier_filter)
    
    min_score = request.GET.get('min_score', '').strip()
    if min_score:
        try:
            min_score_int = int(min_score)
            filtered_ids = []
            for eval in evaluations:
                if eval.get_total_score() >= min_score_int:
                    filtered_ids.append(eval.id)
            evaluations = evaluations.filter(id__in=filtered_ids)
        except ValueError:
            pass
    
    date_from = request.GET.get('date_from', '').strip()
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            evaluations = evaluations.filter(date_evaluation__date__gte=date_from_obj)
        except ValueError:
            pass
    
    date_to = request.GET.get('date_to', '').strip()
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            evaluations = evaluations.filter(date_evaluation__date__lte=date_to_obj)
        except ValueError:
            pass
    
    # Étape 1: Créer un cache pour les premières occurrences de chaque bon de commande
    orders_needed = {eval.bon_commande.numero for eval in evaluations}
    
    all_files = []
    for evaluation in evaluations:
        for fichier in evaluation.bon_commande.fichiers.all():
            if fichier not in all_files:
                all_files.append(fichier)
    all_files.sort(key=lambda f: f.date_importation)
    
    first_occurrence_cache = {}
    for fichier in all_files:
        lignes = list(fichier.lignes.all())
        lignes.sort(key=lambda l: l.numero_ligne)
        for ligne in lignes:
            contenu = ligne.contenu or {}
            order_key = find_order_key(contenu)
            if not order_key:
                continue
            order_val = str(contenu.get(order_key, '')).strip()
            if not order_val:
                continue
            if order_val in orders_needed and order_val not in first_occurrence_cache:
                first_occurrence_cache[order_val] = contenu
                if len(first_occurrence_cache) == len(orders_needed):
                    break
        if len(first_occurrence_cache) == len(orders_needed):
            break
    
    # Étape 2: Construire les données
    data = []
    
    for evaluation in evaluations:
        bon = evaluation.bon_commande
        supplier = evaluation.supplier
        
        # Récupérer la première occurrence de ce bon de commande
        premiere_occurrence = first_occurrence_cache.get(bon.numero)
        
        if premiere_occurrence:
            # Extraire les informations avec get_value_tolerant
            item = get_value_tolerant(premiere_occurrence, 
                exact_candidates=['Order Description', 'Item Description', 'Description'],
                tokens=['order', 'description']) or 'N/A'
            
            pm = get_value_tolerant(premiere_occurrence,
                exact_candidates=['Project Manager', 'PM', 'Manager'],
                tokens=['project', 'manager']) or 'N/A'
            
            pip_end_date_raw = get_value_tolerant(premiere_occurrence,
                exact_candidates=['PIP END DATE', 'PIP End Date', 'Pip End Date'],
                tokens=['pip', 'end', 'date']) or ''
            
            actual_end_date_raw = get_value_tolerant(premiere_occurrence,
                exact_candidates=['ACTUAL END DATE', 'Actual End Date', 'Real End Date'],
                tokens=['actual', 'end', 'date']) or ''
            
            # Normaliser les dates au format YYYY-MM-DD (comme dans PO Progress)
            def normalize_date(date_val):
                """Normalise une date au format YYYY-MM-DD"""
                if not date_val:
                    return ''
                
                date_str = str(date_val).strip()
                
                # Si déjà au format ISO avec timestamp, extraire la date
                if len(date_str) >= 10 and date_str[4] == '-' and date_str[7] == '-':
                    return date_str[:10]  # Prendre seulement YYYY-MM-DD
                
                # Essayer de parser d'autres formats
                date_formats = ['%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d']
                for fmt in date_formats:
                    try:
                        parsed = datetime.strptime(date_str, fmt)
                        return parsed.strftime('%Y-%m-%d')
                    except ValueError:
                        continue
                
                # Par défaut, retourner tel quel
                return date_str
            
            pip_end_date = normalize_date(pip_end_date_raw)
            actual_end_date = normalize_date(actual_end_date_raw)
            
            # Calculer le nombre de jours de retard (supporte plusieurs formats)
            total_days_late = 0
            if pip_end_date and actual_end_date:
                try:
                    # Convertir en string et nettoyer
                    pip_end_str = str(pip_end_date).strip()
                    actual_end_str = str(actual_end_date).strip()
                    
                    # Liste des formats supportés
                    date_formats = [
                        '%Y-%m-%d %H:%M:%S',  # 2025-07-30 00:00:00
                        '%Y-%m-%d',           # 2025-07-30
                        '%d/%m/%Y',           # 30/07/2025
                        '%d/%m/%Y %H:%M:%S'   # 30/07/2025 00:00:00
                    ]
                    
                    pip_end = None
                    actual_end = None
                    
                    # Essayer chaque format pour PIP END DATE
                    for fmt in date_formats:
                        try:
                            pip_end = datetime.strptime(pip_end_str, fmt)
                            break
                        except ValueError:
                            continue
                    
                    # Essayer chaque format pour ACTUAL END DATE
                    for fmt in date_formats:
                        try:
                            actual_end = datetime.strptime(actual_end_str, fmt)
                            break
                        except ValueError:
                            continue
                    
                    if pip_end and actual_end:
                        total_days_late = max(0, (actual_end - pip_end).days)
                    else:
                        total_days_late = 0
                except Exception as e:
                    total_days_late = 0
        else:
            item = 'N/A'
            pm = 'N/A'
            pip_end_date = ''
            actual_end_date = ''
            total_days_late = 0
        
        # Récupérer le montant total du PO
        montant_total = bon.montant_total() if callable(getattr(bon, 'montant_total', None)) else Decimal('0')
        
        # Construire la ligne de données
        row = {
            'N° PO': bon.numero,
            'FOURNISSEURS': supplier,
            'ITEM': str(item),
            'MONTANT PO': float(montant_total),
            'PM': str(pm),
            'Conformité de la commande (quantité & qualité)': evaluation.delivery_compliance,
            'Délai de livraison, implémentation, exécution,…': evaluation.delivery_timeline,
            'Capacité à conseiller (transfert de compétence, formation, Autonomie du client)': evaluation.advising_capability,
            'Qualité du SAV': evaluation.after_sales_qos,
            'Contact relationnel': evaluation.vendor_relationship,
            'NOTE GLOBALE': float(evaluation.vendor_final_rating),
            'DATE DE FIN CONTRACTUELLE(PIP)': pip_end_date,
            'DATE DE FIN REELLE': actual_end_date,
            'NOMBRE TOTAL DE JOURS DE RETARD': total_days_late,
            'PART MTN': getattr(bon.timeline_delay, 'delay_part_mtn', 0) if hasattr(bon, 'timeline_delay') else 0,
            'PART FORCE MAJEURE': getattr(bon.timeline_delay, 'delay_part_force_majeure', 0) if hasattr(bon, 'timeline_delay') else 0,
            'PART FOURNISSEUR': getattr(bon.timeline_delay, 'delay_part_vendor', 0) if hasattr(bon, 'timeline_delay') else 0,
            'Montant Rétention Timeline': float(getattr(bon.timeline_delay, 'retention_amount_timeline', 0)) if hasattr(bon, 'timeline_delay') else 0,
            'Taux Rétention Timeline (%)': f"{float(getattr(bon.timeline_delay, 'retention_rate_timeline', 0)) if hasattr(bon, 'timeline_delay') else 0:.2f}%".replace('.', ','),
        }
        
        data.append(row)
    
    # Créer le DataFrame
    df = pd.DataFrame(data)
    
    # Remplacer les valeurs vides par des chaînes vides pour éviter NaT (seulement si les colonnes existent)
    if 'DATE DE FIN CONTRACTUELLE(PIP)' in df.columns:
        df['DATE DE FIN CONTRACTUELLE(PIP)'] = df['DATE DE FIN CONTRACTUELLE(PIP)'].fillna('')
    if 'DATE DE FIN REELLE' in df.columns:
        df['DATE DE FIN REELLE'] = df['DATE DE FIN REELLE'].fillna('')
    
    # Créer le fichier Excel
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Vendor Evaluations')
        
        # Récupérer le workbook et la feuille
        workbook = writer.book
        worksheet = writer.sheets['Vendor Evaluations']
        
        # Styles
        header_font = Font(bold=True, color="000000", size=11)
        header_fill = PatternFill(start_color="FFCC00", end_color="FFCC00", fill_type="solid")
        alignment_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        alignment_left = Alignment(horizontal="left", vertical="center")
        alignment_right = Alignment(horizontal="right", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Formater l'en-tête
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment_center
            cell.border = thin_border
            
            # Ajuster la largeur des colonnes
            if 'DATE' in column_title or 'N° PO' in column_title:
                worksheet.column_dimensions[get_column_letter(col_num)].width = 20
            elif 'MONTANT' in column_title or 'Rétention' in column_title:
                worksheet.column_dimensions[get_column_letter(col_num)].width = 18
            elif 'NOTE' in column_title:
                worksheet.column_dimensions[get_column_letter(col_num)].width = 15
            elif len(column_title) > 40:
                worksheet.column_dimensions[get_column_letter(col_num)].width = 35
            else:
                worksheet.column_dimensions[get_column_letter(col_num)].width = max(len(column_title) + 2, 15)

        # Formater les colonnes de dates en texte (pour préserver le format YYYY-MM-DD)
        date_cols = ['DATE DE FIN CONTRACTUELLE(PIP)', 'DATE DE FIN REELLE']
        for col_name in date_cols:
            if col_name in df.columns:
                col_idx = df.columns.get_loc(col_name) + 1
                for row_idx in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.number_format = '@'  # Format texte
        
        # Formater les cellules de données
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for idx, cell in enumerate(row):
                cell.border = thin_border
                
                # Alignement selon le type de données
                column_name = df.columns[idx]
                if 'MONTANT' in column_name or 'Rétention' in column_name or 'NOTE' in column_name or 'JOURS' in column_name:
                    cell.alignment = alignment_right
                    if 'MONTANT' in column_name or 'Rétention' in column_name:
                        cell.number_format = '#,##0.00'
                    elif 'NOTE' in column_name:
                        cell.number_format = '0.00'
                elif any(x in column_name for x in ['Conformité', 'Délai', 'Capacité', 'Qualité', 'Contact']):
                    cell.alignment = alignment_center
                elif 'DATE' in column_name:
                    cell.alignment = alignment_center
                else:
                    cell.alignment = alignment_left
    
    # Préparer la réponse HTTP
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"vendor_evaluations_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

@login_required
def export_vendor_ranking(request):
    """
    Export Excel optimisé pour le Vendor Ranking
    Une ligne par PO unique avec toutes les informations demandées
    """
    try:
        from .models import VendorEvaluation, NumeroBonCommande, LigneFichier
        from datetime import datetime
        import pandas as pd
        from decimal import Decimal
        from django.http import HttpResponse
        
        # Récupérer uniquement les POs qui ont des évaluations
        evaluations = VendorEvaluation.objects.select_related('bon_commande', 'evaluator').all()
        
        # Filtrer les évaluations par POs autorisés (Option A) pour les non-superusers
        if not request.user.is_superuser:
            from .views import filter_bons_by_user_service
            allowed_bons = filter_bons_by_user_service(NumeroBonCommande.objects.all(), request.user)
            evaluations = evaluations.filter(bon_commande__in=allowed_bons)
        
        # Extraire les bons de commande qui ont des évaluations
        bons_commande_ids = evaluations.values_list('bon_commande_id', flat=True).distinct()
        bons_commande = NumeroBonCommande.objects.filter(id__in=bons_commande_ids).prefetch_related(
            'fichiers__lignes'
        )
        
        if not bons_commande.exists():
            messages.warning(request, "Aucun bon de commande avec évaluation disponible pour l'export.")
            return redirect('orders:vendor_ranking')
        
        # Créer un dictionnaire des évaluations par bon de commande
        evaluations_dict = {}
        for evaluation in evaluations:
            evaluations_dict[evaluation.bon_commande.numero] = evaluation
        
        # Fonction helper pour trouver la clé "Order" dans le contenu
        def find_order_key(contenu: dict):
            # 1) Cas exact 'Order'
            if 'Order' in contenu:
                return 'Order'
            # 2) Recherche tolérante (espaces/underscores/accents)
            for k in contenu.keys():
                if not k:
                    continue
                norm = ' '.join(str(k).strip().lower().replace('_', ' ').split())
                if 'order' in norm or 'commande' in norm or 'bon' in norm or norm == 'bc':
                    return k
            return None
        
        # Fonction helper pour récupérer une valeur avec normalisation tolérante
        def get_value_tolerant(contenu: dict, exact_candidates=None, tokens=None):
            if not contenu:
                return None
            def norm(s: str):
                return ' '.join(str(s).strip().lower().replace('_', ' ').replace('-', ' ').split())
            normalized = {norm(k): (k, v) for k, v in contenu.items() if k}
            
            if exact_candidates:
                for cand in exact_candidates:
                    nk = norm(cand)
                    if nk in normalized:
                        return normalized[nk][1]
            
            if tokens:
                needed = [norm(t) for t in tokens]
                for nk, (_ok, v) in normalized.items():
                    if all(t in nk for t in needed):
                        return v
            return None
        
        # Étape 1: Créer un cache pour les premières occurrences de chaque bon de commande en une seule passe globale
        orders_needed = {bon.numero for bon in bons_commande}
        
        # Lister tous les fichiers associés, triés par date d'importation (du plus ancien au plus récent)
        all_files = []
        for bon in bons_commande:
            for fichier in bon.fichiers.all():
                if fichier not in all_files:
                    all_files.append(fichier)
        all_files.sort(key=lambda f: f.date_importation)
        
        # Construire le cache des premières occurrences
        first_occurrence_cache = {}
        for fichier in all_files:
            lignes = list(fichier.lignes.all())
            lignes.sort(key=lambda l: l.numero_ligne)
            for ligne in lignes:
                contenu = ligne.contenu or {}
                order_key = find_order_key(contenu)
                if not order_key:
                    continue
                order_val = str(contenu.get(order_key, '')).strip()
                if not order_val:
                    continue
                # Ajouter au cache seulement si pas encore présent (première occurrence)
                if order_val in orders_needed and order_val not in first_occurrence_cache:
                    first_occurrence_cache[order_val] = contenu
                    # Early break: si on a trouvé toutes les commandes, arrêter
                    if len(first_occurrence_cache) == len(orders_needed):
                        break
            # Early break au niveau fichier aussi
            if len(first_occurrence_cache) == len(orders_needed):
                break
        
        # Étape 2: Construire les données pour TOUS les bons de commande
        data = []
        
        for bon in bons_commande:
            # Récupérer l'évaluation si elle existe
            evaluation = evaluations_dict.get(bon.numero)
            supplier = evaluation.supplier if evaluation else 'N/A'
            
            # Récupérer la première occurrence de ce bon de commande
            premiere_occurrence = first_occurrence_cache.get(bon.numero)
            
            if premiere_occurrence:
                # Extraire les informations
                order_status = get_value_tolerant(premiere_occurrence,
                    exact_candidates=['Order Status', 'Status', 'Statut'],
                    tokens=['order', 'status']) or 'N/A'
                
                # Utiliser .get() direct pour Closed Date (comme dans le fichier)
                closed_date = premiere_occurrence.get('Closed Date') or ''
                
                project_manager = get_value_tolerant(premiere_occurrence,
                    exact_candidates=['Project Manager', 'PM', 'Manager'],
                    tokens=['project', 'manager']) or 'N/A'
                
                buyer = get_value_tolerant(premiere_occurrence,
                    exact_candidates=['Buyer', 'Acheteur', 'Purchaser'],
                    tokens=['buyer']) or 'N/A'
                
                order_description = get_value_tolerant(premiere_occurrence,
                    exact_candidates=['Order Description', 'Item Description', 'Description'],
                    tokens=['order', 'description']) or 'N/A'
                
                # Utiliser .get() direct pour Currency et Conversion Rate (comme export_po_progress)
                currency = premiere_occurrence.get('Currency') or 'XOF'
                
                # Utiliser bon.montant_total() comme dans export_po_progress
                po_amount_value = float(bon.montant_total()) if callable(getattr(bon, 'montant_total', None)) else 0.0
                
                # Calculer le montant en XOF (exactement comme dans export_po_progress)
                if currency == 'XOF':
                    exchange_rate = 1.0
                else:
                    # Récupérer le taux de conversion de la première occurrence
                    exchange_rate_str = premiere_occurrence.get('Conversion Rate')
                    try:
                        exchange_rate = float(exchange_rate_str) if exchange_rate_str else 1.0
                    except (TypeError, ValueError):
                        exchange_rate = 1.0
                
                po_amount_xof = po_amount_value * exchange_rate
            else:
                order_status = 'N/A'
                closed_date = ''
                project_manager = 'N/A'
                buyer = 'N/A'
                order_description = 'N/A'
                currency = 'XOF'
                po_amount_value = 0
                po_amount_xof = 0
            
            # Construire la ligne de données
            row = {
                'Purchase Order': bon.numero,
                'Statut PO': str(order_status),
                'Date de fermeture PO': str(closed_date),
                'Demandeur': str(project_manager),
                'Acheteur': str(buyer),
                'Description de la commande': str(order_description),
                'Devise': str(currency),
                'Montant en devise': float(po_amount_value),
                'Montant en XOF': float(po_amount_xof),
                'Conformité de la commande (quantité & qualité)': evaluation.delivery_compliance if evaluation else 0,
                'Délai de livraison, implémentation, exécution': evaluation.delivery_timeline if evaluation else 0,
                'Capacité à conseiller (transfert de compétence, formation, autonomie du client)': evaluation.advising_capability if evaluation else 0,
                'Qualité du SAV': evaluation.after_sales_qos if evaluation else 0,
                'Contact relationnel': evaluation.vendor_relationship if evaluation else 0,
                'Moyenne évaluation demandeur': float(evaluation.vendor_final_rating) if evaluation else 0,
                'Observation': '',  # Colonne vide pour observations manuelles
            }
            
            data.append(row)
        
        # Créer le DataFrame
        df = pd.DataFrame(data)
        
        # Remplacer les valeurs vides
        df['Date de fermeture PO'] = df['Date de fermeture PO'].fillna('')
        
        # Créer la réponse HTTP avec le fichier Excel
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename="vendor_ranking_export_{timestamp}.xlsx"'
        
        # Écrire le DataFrame dans le fichier Excel
        with pd.ExcelWriter(response, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Vendor Ranking')
            
            # Formater le fichier Excel
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            workbook = writer.book
            worksheet = writer.sheets['Vendor Ranking']
            
            # Styles (couleurs MTN: Noir et Jaune)
            header_font = Font(bold=True, color="000000", size=11)
            header_fill = PatternFill(start_color="FFCC00", end_color="FFCC00", fill_type="solid")
            alignment_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
            alignment_left = Alignment(horizontal="left", vertical="center")
            alignment_right = Alignment(horizontal="right", vertical="center")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Appliquer le style aux en-têtes
            for col_idx, col_name in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment_center
                cell.border = thin_border
            
            # Appliquer les styles et bordures aux cellules de données
            for row_idx in range(2, len(df) + 2):
                for col_idx, col_name in enumerate(df.columns, 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.border = thin_border
                    
                    # Alignement selon le type de colonne
                    if col_name in ['Purchase Order', 'Statut PO', 'Demandeur', 
                                    'Acheteur', 'Description de la commande', 'Devise', 'Observation']:
                        cell.alignment = alignment_left
                    elif col_name in ['Montant en devise', 'Montant en XOF',
                                      'Conformité de la commande (quantité & qualité)',
                                      'Délai de livraison, implémentation, exécution',
                                      'Capacité à conseiller (transfert de compétence, formation, autonomie du client)',
                                      'Qualité du SAV', 'Contact relationnel', 'Moyenne évaluation demandeur']:
                        cell.alignment = alignment_right
                    else:
                        cell.alignment = alignment_center
            
            # Ajuster la largeur des colonnes
            for idx, col in enumerate(df.columns, 1):
                max_length = max(
                    df[col].astype(str).apply(len).max(),
                    len(col)
                ) + 2
                column_letter = get_column_letter(idx)
                worksheet.column_dimensions[column_letter].width = min(max_length, 50)
        
        return response
        
    except Exception as e:
        logger.error(f"Erreur lors de l'export Vendor Ranking: {str(e)}")
        messages.error(request, f"Erreur lors de l'export: {str(e)}")
        return redirect('orders:vendor_ranking')


@login_required
def export_fichier_complet(request, fichier_id):
    """
    Exporte toutes les données d'un fichier importé en Excel avec les données mises à jour des réceptions.
    Cette fonction réutilise la logique de fusion des données de export_bon_excel mais sans filtrer par numéro de commande.
    """
    from .models import FichierImporte, Reception
    
    try:
        # Récupérer le fichier importé
        fichier = get_object_or_404(FichierImporte, id=fichier_id)
        
        # Récupérer toutes les lignes du fichier avec leur ID
        lignes_fichier = fichier.lignes.all().order_by('numero_ligne')
        
        # Ajouter une fonction utilitaire pour extraire les valeurs numériques
        def extract_numeric_values(data):
            numeric_fields = ['Ordered Quantity', 'Quantity Delivered', 'Quantity Not Delivered', 'Price']
            for field in numeric_fields:
                if field in data:
                    try:
                        # Convertir en float puis en int si possible
                        value = data[field]
                        if isinstance(value, str):
                            value = value.replace(',', '').strip()
                        num = float(value) if value not in (None, '') else 0.0
                        if num.is_integer():
                            data[field] = int(num)
                        else:
                            data[field] = num
                    except (TypeError, ValueError):
                        data[field] = 0
            return data
        
        # Créer une liste des données avec l'ID de la ligne
        contenu_data = []
        for ligne in lignes_fichier:
            data = ligne.contenu.copy()
            data['_business_id'] = ligne.business_id  # Store business ID instead of row index
            data = extract_numeric_values(data)  # Extraire les valeurs numériques
            contenu_data.append(data)
        
        # Initialiser le dictionnaire des réceptions indexé par business_id
        receptions = {}
        
        # Récupérer toutes les réceptions pour ce fichier
        receptions_queryset = Reception.objects.filter(
            fichier=fichier
        ).select_related('fichier', 'bon_commande').order_by('business_id')
        
        # Convertir les réceptions en dictionnaire indexé par business_id
        for reception in receptions_queryset:
            receptions[str(reception.business_id)] = {
                'quantity_delivered': reception.quantity_delivered,
                'ordered_quantity': reception.ordered_quantity,
                'quantity_not_delivered': reception.quantity_not_delivered,
                'amount_delivered': reception.amount_delivered,
                'quantity_payable': reception.quantity_payable,
                'amount_payable': reception.amount_payable  # Use the desired header
            }
        
        # Ajouter les données de réception aux données brutes en utilisant l'ID de ligne
        if contenu_data and isinstance(contenu_data, list):
            for item in contenu_data:
                # Utiliser l'ID de la ligne pour la correspondance
                idx = item.get('_business_id')
                
                if idx is not None and str(idx) in receptions:
                    # Si on a une réception pour cette ligne, utiliser ses valeurs
                    rec = receptions[str(idx)]
                    # Ajouter les données de réception
                    item['Quantity Delivered'] = rec['quantity_delivered']
                    item['Ordered Quantity'] = rec['ordered_quantity']
                    item['Quantity Not Delivered'] = rec['quantity_not_delivered']
                    item['Amount Delivered'] = rec['amount_delivered']
                    item['Quantity Payable'] = rec['quantity_payable']
                    item['Amount Payable'] = rec['amount_payable']
                elif 'Ordered Quantity' in item:
                    # Initialiser avec des valeurs par défaut si pas de réception existante
                    try:
                        ordered_qty = float(item['Ordered Quantity']) if item['Ordered Quantity'] is not None else 0
                        # Initialiser les valeurs par défaut
                        item['Quantity Delivered'] = 0
                        item['Quantity Not Delivered'] = ordered_qty
                        item['Amount Delivered'] = 0
                        item['Quantity Payable'] = 0
                        item['Amount Payable'] = 0
                    except (ValueError, TypeError):
                        item['Quantity Delivered'] = 0
                        item['Quantity Not Delivered'] = 0
                        item['Amount Delivered'] = 0
                        item['Quantity Payable'] = 0
                        item['Amount Payable'] = 0
                        
                        # Conserver la colonne Quantity Delivered
        
        # Supprimer les clés techniques qui ne doivent pas apparaître dans l'export
        for item in contenu_data:
            if '_business_id' in item:
                del item['_business_id']
        
        # Générer le fichier Excel
        import pandas as pd
        from django.http import HttpResponse
        from io import BytesIO
        
        # Créer un DataFrame pandas avec les données
        df = pd.DataFrame(contenu_data)
        
        # Créer un buffer en mémoire pour stocker le fichier Excel
        output = BytesIO()
        
        # Déterminer le nom du fichier
        filename = f"Fichier_complet_{fichier.id}_updated.xlsx"
        
        # Créer un writer Excel
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Premier onglet avec les données principales
            df.to_excel(writer, index=False, sheet_name='Données')
            
            # Deuxième onglet avec les informations sur le fichier
            info_data = {
                'Information': ['Nom du fichier', 'Date d\'importation', 'Nombre de lignes'],
                'Valeur': [
                    os.path.basename(fichier.fichier.name),
                    fichier.date_importation.strftime('%Y-%m-%d %H:%M'),
                    fichier.nombre_lignes
                ]
            }
            info_df = pd.DataFrame(info_data)
            info_df.to_excel(writer, index=False, sheet_name='Informations')
        
        # Configurer la réponse HTTP
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except Exception as e:
        messages.error(request, f"Erreur lors de l'exportation du fichier : {str(e)}")
        return redirect('admin:orders_fichierimporte_changelist')


@login_required
def export_bon_excel(request, bon_id):
    """
    Exporte les données d'un bon de commande en Excel avec les données mises à jour des réceptions.
    Cette fonction réutilise la logique de fusion des données de details_bon.
    """
    from .models import NumeroBonCommande, FichierImporte, Reception
    
    selected_order_number = request.GET.get('selected_order_number')

    if bon_id == 'search' and 'order_number' in request.GET:
        selected_order_number = request.GET.get('order_number')
        try:
            bon_commande = NumeroBonCommande.objects.get(numero=selected_order_number)
            fichier = bon_commande.fichiers.order_by('-date_importation').first()
            if not fichier:
                messages.warning(request, f'Purchase order {selected_order_number} exists but is not associated with any file.')
                return redirect('orders:accueil')
        except NumeroBonCommande.DoesNotExist:
            messages.error(request, f'Purchase order {selected_order_number} not found.')
            return redirect('orders:accueil')
    else:
        fichier = get_object_or_404(FichierImporte, id=bon_id)
        if not selected_order_number and fichier.bons_commande.exists():
            selected_order_number = fichier.bons_commande.first().numero
    
    # Récupérer les données depuis le modèle LigneFichier
    try:
        # Récupérer toutes les lignes du fichier avec leur ID
        lignes_fichier = fichier.lignes.all().order_by('numero_ligne')
        
        # Ajouter une fonction utilitaire pour extraire les valeurs numériques
        def extract_numeric_values(data):
            numeric_fields = ['Ordered Quantity','Quantity Delivered', 'Quantity Not Delivered', 'Price']
            for field in numeric_fields:
                if field in data:
                    try:
                        # Convertir en float puis en int si possible
                        value = data[field]
                        if isinstance(value, str):
                            value = value.replace(',', '').strip()
                        num = float(value) if value not in (None, '') else 0.0
                        if num.is_integer():
                            data[field] = int(num)
                        else:
                            data[field] = num
                    except (TypeError, ValueError):
                        data[field] = 0
            return data
        
        # Créer une liste des données avec l'ID de la ligne
        contenu_data = []
        for ligne in lignes_fichier:
            data = ligne.contenu.copy()
            data['_business_id'] = ligne.business_id  # Store business ID instead of row index
            data = extract_numeric_values(data)  # Extraire les valeurs numériques
            contenu_data.append(data)
        
        # Initialiser le dictionnaire des réceptions indexé par business_id
        receptions = {}
        
        # Variables pour les informations financières
        montant_total = 0
        montant_recu = 0
        taux_avancement = 0
        currency = 'N/A'
        
        # Si un numéro de bon de commande est sélectionné, charger les réceptions existantes
        if selected_order_number:
            try:
                # Récupérer le bon de commande
                bon_commande = NumeroBonCommande.objects.get(numero=selected_order_number)
                
                # Récupérer les réceptions pour ce bon de commande et ce fichier
                receptions_queryset = Reception.objects.filter(
                    bon_commande=bon_commande,
                    fichier=fichier
                ).select_related('fichier', 'bon_commande').order_by('business_id')
                
                # Convertir les réceptions en dictionnaire indexé par business_id
                for reception in receptions_queryset:
                    receptions[str(reception.business_id)] = {
                        'quantity_delivered': reception.quantity_delivered,
                        'ordered_quantity': reception.ordered_quantity,
                        'quantity_not_delivered': reception.quantity_not_delivered,
                        'amount_delivered': reception.amount_delivered,
                        'quantity_payable': reception.quantity_payable,
                        'unit_price': reception.unit_price,
                        'amount_payable': reception.amount_payable  # Use the desired header
                    }
                
                # Calculer les montants totaux pour l'onglet d'information
                montant_total = bon_commande.montant_total()
                montant_recu = bon_commande.montant_recu()
                if montant_total > 0:
                    taux_avancement = (montant_recu / montant_total) * 100
                else:
                    taux_avancement = 0
                
                # Récupérer la devise depuis le bon de commande
                currency = bon_commande.get_currency() if callable(getattr(bon_commande, 'get_currency', None)) else 'N/A'
                    
            except NumeroBonCommande.DoesNotExist:
                # Si le bon de commande n'existe pas encore, on continue avec un dictionnaire vide
                pass
                
    except Exception as e:
        messages.error(request, f"Erreur lors de la récupération des données : {str(e)}")
        return redirect('orders:details_bon', bon_id=bon_id)
        
    raw_data = []
    headers = []
    colonne_order = None
    
    if contenu_data:
        # Utiliser les en-têtes stockés dans le fichier importé
        
        # Traitement des données structurées (liste de dictionnaires)
        if isinstance(contenu_data, list) and len(contenu_data) > 0 and isinstance(contenu_data[0], dict):
            # Récupérer tous les en-têtes uniques
            all_keys = set()
            for item in contenu_data:
                all_keys.update(item.keys())
            
            # Trier les en-têtes par ordre alphabétique
            headers = sorted(list(all_keys))
            
            # Debug: afficher les headers disponibles
            print(f"[DEBUG export_bon_excel] Headers disponibles ({len(headers)}): {headers}")
            
            # Trouver la colonne qui contient le numéro de commande
            for cle in headers:
                if cle.upper() in ['ORDER', 'ORDRE', 'BON', 'BON_COMMANDE', 'COMMANDE', 'BC', 'NUM_BC', 'PO', 'PO_NUMBER']:
                    colonne_order = cle
                    break
            
            # Filtrer les données si un numéro de commande est sélectionné
            if selected_order_number and colonne_order:
                raw_data = [
                    item for item in contenu_data 
                    if colonne_order in item and str(item[colonne_order]) == str(selected_order_number)
                ]
            else:
                raw_data = contenu_data
    
    # Ajouter les données de réception aux données brutes en utilisant l'ID de ligne
    if contenu_data and isinstance(contenu_data, list):
        for item in contenu_data:
            # Utiliser l'ID de la ligne pour la correspondance
            idx = item.get('_business_id')
            
            if idx is not None and str(idx) in receptions:
                # Si on a une réception pour cette ligne, utiliser ses valeurs
                rec = receptions[str(idx)]
                
                item['Quantity Delivered'] = rec['quantity_delivered']
                item['Ordered Quantity'] = rec['ordered_quantity']
                item['Quantity Not Delivered'] = rec['quantity_not_delivered']
                item['Amount Delivered'] = rec['amount_delivered']
                item['Quantity Payable'] = rec['quantity_payable']
                item['Amount Payable'] = rec['amount_payable']

            elif 'Ordered Quantity' in item:
                # Initialiser avec des valeurs par défaut si pas de réception existante
                try:
                    ordered_qty = float(item['Ordered Quantity']) if item['Ordered Quantity'] is not None else 0
                    # Utiliser Quantity Delivered au lieu de Quantity Delivered
                    item['Quantity Delivered'] = 0
                    item['Quantity Not Delivered'] = ordered_qty
                    item['Amount Delivered'] = 0
                    item['Quantity Payable'] = 0
                    item['Amount Payable'] = 0
                    

                except (ValueError, TypeError):
                    item['Quantity Delivered'] = 0
                    item['Quantity Not Delivered'] = 0
                    item['Amount Delivered'] = 0
                    item['Quantity Payable'] = 0
                    item['Amount Payable'] = 0                    
    
    # Nettoyer les données avant export: retirer lignes d'erreur et totalement vides
    filtered_raw = []
    for item in raw_data:
        if not isinstance(item, dict):
            continue
        # Exclure toute ligne d'erreur explicite
        if 'error' in item:
            continue
        # Considérer vides les lignes dont toutes les valeurs non techniques sont vides/None
        non_tech_keys = [k for k in item.keys() if not k.startswith('_')]
        has_value = any(
            (item.get(k) is not None and str(item.get(k)).strip() != '')
            for k in non_tech_keys
        )
        if not has_value:
            continue
        filtered_raw.append(item)
    raw_data = filtered_raw

    # Supprimer les clés techniques qui ne doivent pas apparaître dans l'export
    for item in raw_data:
        if '_business_id' in item:
            del item['_business_id']
    
    # Générer le fichier Excel
    import pandas as pd
    from django.http import HttpResponse
    from io import BytesIO
    
    # Créer un DataFrame pandas avec les données
    df = pd.DataFrame(raw_data)
    
    # Créer un buffer en mémoire pour stocker le fichier Excel
    output = BytesIO()
    
    # Déterminer le nom du fichier
    filename = f"PO_{selected_order_number}_updated.xlsx" if selected_order_number else f"Fichier_{fichier.id}_updated.xlsx"
    
    # Créer un writer Excel
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Premier onglet avec les données principales
        df.to_excel(writer, index=False, sheet_name='Données')
        
        # Deuxième onglet avec les informations financières
        if selected_order_number:
            # Créer un DataFrame pour les informations financières
            info_data = {
                'Information': ['Devise', 'Montant Total', 'Montant Total Reçu', 'Taux d\'Avancement'],
                'Valeur': [
                    currency,
                    f"{montant_total:,.2f}",
                    f"{montant_recu:,.2f}",
                    f"{taux_avancement:.2f}%"
                ]
            }
            info_df = pd.DataFrame(info_data)
            info_df.to_excel(writer, index=False, sheet_name='Informations')
    
    # Configurer la réponse HTTP
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


def export_msrn_selection_to_excel(queryset):
    """
    Export des rapports MSRN sélectionnés pour l'admin,
    avec détails des signatures (nom + statut emoji).
    """
    import io
    from datetime import datetime
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "MSRN Reports"
    
    # Styles
    header_font = Font(bold=True, color="yellow")
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    
    # En-têtes
    headers = [
        "Report Number", "PO Number", "Supplier", "Company", "Montant Reçu", 
        "Currency", "Retention %", "Progress %", "CPU",
        "Project Manager", "Coordinator", "Senior PM", "Senior Technical Lead",
        "Manager Portfolio", "GM EPMO", "Vendor", "Created At"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Mapping rôle -> nom affichage pour les signatures
    role_map = {
        'project_manager': 'Project Manager',
        'project_coordinator': 'Coordinator',
        'senior_pm': 'Senior PM',
        'senior_technical_lead': 'Senior Technical Lead',
        'manager_portfolio': 'Manager Portfolio',
        'gm_epmo': 'GM EPMO',
        'vendor': 'Vendor'
    }

    # Données
    # Utilisation de select_related pour le bon_commande et prefetch_related pour les signatures
    for row_num, report in enumerate(queryset.select_related('bon_commande').prefetch_related('signature_tracking'), 2):
        # Initialiser les signatures avec '-'
        sigs_status = {v: '-' for v in role_map.values()}
        for sig in report.signature_tracking.all():
            if sig.signatory_role in role_map:
                display_role = role_map[sig.signatory_role]
                emoji = "✅" if sig.date_received else "⏳"
                sigs_status[display_role] = f"{sig.signatory_name} {emoji}"
                
        data = [
            report.report_number,
            report.bon_commande.numero if report.bon_commande else '-',
            report.supplier_snapshot or '-',
            'MTN',
            float(report.montant_recu_snapshot or 0),
            report.currency_snapshot or 'XOF',
            float(report.retention_rate or 0),
            float(report.progress_rate_snapshot or 0),
            report.cpu_snapshot or '-',
            sigs_status['Project Manager'],
            sigs_status['Coordinator'],
            sigs_status['Senior PM'],
            sigs_status['Senior Technical Lead'],
            sigs_status['Manager Portfolio'],
            sigs_status['GM EPMO'],
            sigs_status['Vendor'],
            report.created_at.strftime('%Y-%m-%d %H:%M') if report.created_at else '-',
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            
            # Format pour les montants
            if col_num == 5:  # Montant Reçu
                cell.number_format = '#,##0'
            elif col_num in (7, 8):  # Taux
                cell.number_format = '0.00'
    
    # Ajuster la largeur des colonnes (17 colonnes totales)
    column_widths = [15, 15, 25, 10, 18, 8, 15, 15, 12, 25, 25, 25, 25, 25, 25, 25, 18]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Figer la première ligne
    ws.freeze_panes = 'A2'
    
    # Créer le buffer de retour
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Créer la réponse HTTP
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"MSRN_Reports_Selection_{timestamp}.xlsx"
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response
