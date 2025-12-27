import os
import pandas as pd
from decimal import Decimal
from django.utils import timezone
from .models import (
    FichierImporte, LigneFichier, Reception, 
    NumeroBonCommande, InitialReceptionBusiness
)
from .utils import round_decimal, normalize_business_id

def process_chunk(records, start_index, fichier_importe, po_cpu_map):
    """
    Traite un lot d'enregistrements (records) pour l'import.
    
    Args:
        records (list[dict]): Liste des lignes (dicts) à traiter.
        start_index (int): Index de départ pour le numéro de ligne.
        fichier_importe (FichierImporte): L'instance du fichier parent.
        po_cpu_map (dict): Dictionnaire partagé pour stocker les CPUs par PO.
        
    Returns:
        int: Nombre de lignes traitées (toujours len(records)).
    """
    user = fichier_importe.utilisateur
    
    # 1. Identifier les business_id dans ce chunk pour vérifier l'existence
    chunk_business_ids = []
    temp_records_map = {} # map business_id -> record
    
    for i, record in enumerate(records):
        if not isinstance(record, dict): continue
        temp_lf = LigneFichier(contenu=record)
        bid = temp_lf.generate_business_id()
        if bid:
            chunk_business_ids.append(bid)
            temp_records_map[bid] = (i, record)
            
    # 2. Trouver les LigneFichier existantes avec ces business_id
    existing_lignes = LigneFichier.objects.filter(business_id__in=chunk_business_ids)
    existing_lignes_map = {lf.business_id: lf for lf in existing_lignes}
    
    lignes_to_create = []
    lignes_to_update = []
    business_ids = set()
    chunk_reception_data = [] # Stores tuple (business_id, record, clean_values)
    
    # Helper pour scanner les valeurs
    def scan_value(d, must_include):
        for k, v in d.items():
            if not k: continue
            kl = k.lower()
            if all(m in kl for m in must_include):
                return v
        return None
    
    # Traitement principal
    for i, record in enumerate(records):
        if not isinstance(record, dict): continue
        
        line_num = start_index + i + 1
        
        # Génération business_id (simulation)
        temp_lf = LigneFichier(contenu=record)
        bid = temp_lf.generate_business_id()
        
        # Si une ligne existe déjà avec ce business_id, on la met à jour
        # pour qu'elle pointe vers le NOUVEAU fichier (réconciliation)
        if bid and bid in existing_lignes_map:
            existing_lf = existing_lignes_map[bid]
            existing_lf.fichier = fichier_importe
            existing_lf.numero_ligne = line_num # Mettre à jour la position
            # On pourrait aussi mettre à jour le contenu si besoin, mais supposons qu'il est similaire
            existing_lf.contenu = record 
            lignes_to_update.append(existing_lf)
        else:
            # Sinon, on crée une nouvelle ligne
            lf = LigneFichier(
                fichier=fichier_importe,
                numero_ligne=line_num,
                contenu=record,
                business_id=bid
            )
            lignes_to_create.append(lf)
            
        if bid:
            business_ids.add(bid)
            
            # Extraction données pour Reception (reste inchangé)
            # Colonnes clés (basé sur FichierImporte.save logic)
            order_key = None
            for k in record.keys():
                if not k: continue
                kl = k.lower()
                if 'order' in kl or 'commande' in kl or 'bon' in kl or 'bc' in kl:
                    order_key = k
                    break
            
            if order_key and record.get(order_key):
                order_number = str(record[order_key]).strip()
                if order_number.lower() not in ('', 'false', 'true', 'none', 'null', 'nan', '0'):
                    # Valeurs numériques
                    rv = record.get('Received Quantity')
                    if rv in (None, ''): rv = scan_value(record, ['received', 'quantity'])
                    received_qty = round_decimal(rv)
                    
                    ov = record.get('Ordered Quantity')
                    if ov in (None, ''): ov = scan_value(record, ['ordered', 'quantity'])
                    ordered_qty = round_decimal(ov)
                    
                    pv = record.get('Price')
                    if pv in (None, ''): pv = scan_value(record, ['price'])
                    unit_price = round_decimal(pv)
                    
                    chunk_reception_data.append({
                        'bid': bid,
                        'order_number': order_number,
                        'received_qty': received_qty,
                        'ordered_qty': ordered_qty,
                        'unit_price': unit_price
                    })
                    
                    # Extraction CPU
                    # Chercher colonne CPU
                    cpu_val = None
                    for k, v in record.items():
                        if k and str(k).strip().upper() == 'CPU' and v:
                            cpu_val = str(v).strip()
                            if ' - ' in cpu_val:
                                cpu_val = cpu_val.split('-')[-1].strip()
                            break
                    if cpu_val and cpu_val.upper() not in ('N/A', 'NA', 'NULL', 'NONE', ''):
                        if order_number not in po_cpu_map: # Premier trouvé gagne
                            po_cpu_map[order_number] = cpu_val

    # Bulk Create LigneFichier
    if lignes_to_create:
        LigneFichier.objects.bulk_create(lignes_to_create, batch_size=2000)
    
    # Bulk Update LigneFichier (Manquant précédemment)
    if lignes_to_update:
        LigneFichier.objects.bulk_update(
            lignes_to_update, 
            ['fichier', 'numero_ligne', 'contenu'], 
            batch_size=2000
        )
    
    # Traitement Réceptions
    if not chunk_reception_data:
        return len(records)

    # Récupérer/Créer les POs en masse pour ce chunk
    chunk_po_nums = {d['order_number'] for d in chunk_reception_data}
    existing_pos = NumeroBonCommande.objects.filter(numero__in=chunk_po_nums)
    po_map = {po.numero: po for po in existing_pos}
    
    new_pos = []
    for num in chunk_po_nums:
        if num not in po_map:
            new_pos.append(NumeroBonCommande(numero=num))
    
    if new_pos:
        NumeroBonCommande.objects.bulk_create(new_pos)
        # Re-fetch pour avoir les IDs
        for po in NumeroBonCommande.objects.filter(numero__in=[p.numero for p in new_pos]):
            po_map[po.numero] = po
            # Associer le fichier (M2M ne supporte pas bulk_create direct facilement, on fera loop)
            po.fichiers.add(fichier_importe)
    
    # Associer fichier aux POs existants aussi
    for po in existing_pos:
            po.fichiers.add(fichier_importe)

    # Charger Réceptions existantes
    chunk_bids = {d['bid'] for d in chunk_reception_data}
    existing_receptions = Reception.objects.filter(business_id__in=chunk_bids).select_related('bon_commande')
    reception_map = {r.business_id: r for r in existing_receptions}
    
    receptions_to_create = []
    receptions_to_update = []
    
    # InitialReceptionBusiness
    irb_existing = InitialReceptionBusiness.objects.filter(business_id__in=chunk_bids)
    irb_map = {i.business_id: i for i in irb_existing}
    irb_to_create = []
    irb_to_update = []

    processed_bids_in_chunk = set()

    for data in chunk_reception_data:
        bid = data['bid']
        if bid in processed_bids_in_chunk: continue
        processed_bids_in_chunk.add(bid)

        po = po_map.get(data['order_number'])
        if not po: continue
        
        ordered = data['ordered_qty']
        received = data['received_qty']
        price = data['unit_price']
        
        # Reception Logic
        rec = reception_map.get(bid)
        if rec:
            # Update logic
            rec.ordered_quantity = ordered
            rec.unit_price = price
            rec.received_quantity = received
            rec.quantity_not_delivered = max(Decimal('0'), ordered - rec.quantity_delivered)
            
            # Recalculer les montants lors de la mise à jour (pour corriger les existants aussi)
            rec.amount_delivered = round_decimal(rec.quantity_delivered * price)
            rec.amount_not_delivered = round_decimal(rec.quantity_not_delivered * price)
            
            # Recalculer le payable (si rétention existe déjà sur le bon, il faudrait la récupérer, 
            # mais ici en bulk on fait simple : recalcul basé sur la quantité livrée actuelle)
            # Pour être précis, il faudrait charger le taux de rétention du PO
            # On assume que quantity_payable suit quantity_delivered (sauf si rétention appliquée, gérée par save())
            # Mais en bulk_update, save() n'est pas appelé. On doit calculer ici.
            
            # Récupérer le taux de rétention (si disponible dans po_map ou via une query séparée optimisée)
            # Pour l'instant, on assume rétention = 0 ou on garde la logique proportionnelle existante
            # Si quantity_payable était déjà set, on peut essayer de conserver le ratio, 
            # mais le plus sûr est de réinitialiser sur la base de delivered si on veut corriger les 0.
            rec.quantity_payable = rec.quantity_delivered # Reset to delivered (retention applied later via UI/Save if needed)
            rec.amount_payable = round_decimal(rec.quantity_payable * price)

            rec.fichier = fichier_importe
            rec.date_modification = timezone.now()
            receptions_to_update.append(rec)
        else:
            # Create logic
            # Nouvelle réception pour ce fichier
            q_not_delivered = max(Decimal('0'), ordered - received)
            
            # Calculer les montants initiaux
            amt_delivered = round_decimal(received * price)
            amt_not_delivered = round_decimal(q_not_delivered * price)
            # Quantity/Amount Payable (initial = 0 ou basé sur received selon logique métier, 
            # ici on initialise proprement pour éviter les NULL/0 injustifiés)
            # Par défaut payable = delivered (si pas de rétention initiale)
            # On assume rétention 0 à la création
            qty_payable = received 
            amt_payable = amt_delivered
            
            new_rec = Reception(
                bon_commande=po,
                fichier=fichier_importe,
                business_id=bid,
                ordered_quantity=ordered,
                received_quantity=received,
                quantity_delivered=received, # Initial delivery = received
                quantity_not_delivered=q_not_delivered,
                unit_price=price,
                amount_delivered=amt_delivered,
                amount_not_delivered=amt_not_delivered,
                quantity_payable=qty_payable,
                amount_payable=amt_payable,
                user=user.email if user and user.email else 'system_import',
                date_modification=timezone.now()
            )
            receptions_to_create.append(new_rec)
            reception_map[bid] = new_rec
        
        # InitialReceptionBusiness Logic
        mt_total = round_decimal(ordered * price)
        mt_recu = round_decimal(received * price)
        taux = Decimal('0')
        if mt_total > 0:
            taux = round_decimal((mt_recu / mt_total) * Decimal('100'))
        
        irb = irb_map.get(bid)
        if irb:
            irb.bon_commande = po
            irb.source_file = fichier_importe
            irb.received_quantity = received
            irb.montant_total_initial = mt_total
            irb.montant_recu_initial = mt_recu
            irb.taux_avancement_initial = taux
            irb_to_update.append(irb)
        else:
            new_irb = InitialReceptionBusiness(
                business_id=bid,
                bon_commande=po,
                source_file=fichier_importe,
                received_quantity=received,
                montant_total_initial=mt_total,
                montant_recu_initial=mt_recu,
                taux_avancement_initial=taux
            )
            irb_to_create.append(new_irb)
            irb_map[bid] = new_irb

    # Executer Bulk Operations
    if receptions_to_create:
        Reception.objects.bulk_create(receptions_to_create, batch_size=2000)
    if receptions_to_update:
        Reception.objects.bulk_update(
            receptions_to_update, 
            [
                'ordered_quantity', 'unit_price', 'received_quantity', 'quantity_not_delivered', 
                'fichier', 'date_modification',
                'amount_delivered', 'amount_not_delivered', 'quantity_payable', 'amount_payable'
            ],
            batch_size=2000
        )
    
    if irb_to_create:
        InitialReceptionBusiness.objects.bulk_create(irb_to_create, batch_size=2000)
    if irb_to_update:
        InitialReceptionBusiness.objects.bulk_update(
            irb_to_update,
            ['bon_commande', 'source_file', 'received_quantity', 'montant_total_initial', 'montant_recu_initial', 'taux_avancement_initial'],
            batch_size=2000
        )
    
    return len(records)

def import_file_optimized(fichier_importe, file_path=None):
    """
    Import optimisé pour gros fichiers (lecture par chunks + bulk operations).
    Peut être appelé par Celery ou par FichierImporte.save().
    
    Si file_path est fourni, l'utilise (ex: temp file de celery).
    Sinon, utilise fichier_importe.fichier.path (si local) ou télécharge si S3.
    """
    # Si pas de path fourni, on essaye de résoudre
    temp_file_created = False
    target_path = file_path
    
    if not target_path:
        try:
            target_path = fichier_importe.fichier.path
        except NotImplementedError:
            # Cas S3/Storage distant -> téléchargement temporaire
            import tempfile
            from django.core.files.storage import default_storage
            
            ext = os.path.splitext(fichier_importe.fichier.name)[1]
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix=ext)
            target_path = tmp.name
            tmp.close()
            
            with fichier_importe.fichier.open('rb') as f:
                with open(target_path, 'wb') as d:
                    d.write(f.read())
            temp_file_created = True

    try:
        # Déterminer extension
        ext = os.path.splitext(target_path)[1].lower().lstrip('.')
        if not ext and fichier_importe.extension:
            ext = fichier_importe.extension
        
        # Variables globales pour le traitement
        total_rows = 0
        po_cpu_map = {}  # PO -> CPU
        
        # NETTOYAGE PREALABLE : Supprimer les anciennes lignes si on rejoue l'import
        # Cela évite l'erreur UniqueViolation sur (fichier_id, numero_ligne)
        if fichier_importe.lignes.exists():
            print(f"Nettoyage des {fichier_importe.lignes.count()} lignes existantes pour le fichier {fichier_importe.id}...")
            fichier_importe.lignes.all().delete()
        
        BATCH_SIZE = 5000
        
        if ext in ['xlsx', 'xlsm', 'xlsb']:
            from openpyxl import load_workbook
            # Read-only mode for memory efficiency
            wb = load_workbook(target_path, read_only=True, data_only=True)
            try:
                ws = wb.active
                headers = None
                chunk = []
                
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i == 0:
                        # Capture headers
                        headers = [str(h).strip() if h is not None else f"col_{j}" for j, h in enumerate(row)]
                        continue
                    
                    # Convert row to dict
                    row_dict = {}
                    has_data = False
                    for j, val in enumerate(row):
                        if j < len(headers):
                            v = str(val) if val is not None else None
                            row_dict[headers[j]] = v
                            if v: has_data = True
                    
                    if has_data:
                        chunk.append(row_dict)
                    
                    if len(chunk) >= BATCH_SIZE:
                        process_chunk(chunk, total_rows, fichier_importe, po_cpu_map)
                        total_rows += len(chunk)
                        chunk = []
                
                if chunk:
                    process_chunk(chunk, total_rows, fichier_importe, po_cpu_map)
                    total_rows += len(chunk)
            finally:
                wb.close()
                
        else:
            # CSV (Pandas)
            # Fallback for others to try csv default
            for chunk_df in pd.read_csv(target_path, chunksize=BATCH_SIZE, dtype=str, keep_default_na=False):
                # Nettoyage NaN -> None
                chunk_df = chunk_df.where(pd.notnull(chunk_df), None)
                records = chunk_df.to_dict('records')
                process_chunk(records, total_rows, fichier_importe, po_cpu_map)
                total_rows += len(records)

        # Mise à jour finale
        fichier_importe.nombre_lignes = total_rows
        # Assurer que l'extension est sauvegardée
        if not fichier_importe.extension:
            fichier_importe.extension = ext
        fichier_importe.save(update_fields=['nombre_lignes', 'extension'])
        
        # Mise à jour CPU sur les POs
        if po_cpu_map:
            pos_to_update = []
            pos = NumeroBonCommande.objects.filter(numero__in=po_cpu_map.keys())
            for po in pos:
                cpu = po_cpu_map.get(po.numero)
                if cpu and po.cpu != cpu:
                    po.cpu = cpu
                    pos_to_update.append(po)
            
            if pos_to_update:
                NumeroBonCommande.objects.bulk_update(pos_to_update, ['cpu'], batch_size=1000)

        return total_rows, len(po_cpu_map)

    finally:
        # Cleanup temp file if we created it
        if temp_file_created and target_path and os.path.exists(target_path):
            try:
                os.remove(target_path)
            except Exception:
                pass
